from functools import wraps
import datetime
import os
import io
import glob
import re
import openpyxl
import pandas as pd


mail_folder = '//192.168.0.21/CustomerService/CSD/Incoming Mail Processing'
excel_folder = os.path.join(mail_folder, 'Mail Opening', '{}', '{}')
export_folder = os.path.join(mail_folder, 'Export')

# Find Excel Files using glob.glob
def find_mail_excels(years=None, months=None):
    # if years and months are None
    year_pattern = '**'
    month_pattern = '*.xl*'

    if years is not None:
      if isinstance(years, int) or isinstance(years, float):
        year_pattern = '{:04d}'.format(int(years))
      else:
        for year in years:
            year_pattern += str(year)

    if months is not None:
      if isinstance(months, int) or isinstance(months, float):
        month_pattern = '*{:02d}*.xl*'.format(int(months))
      else:
        for mth in months:
            month_pattern += str(mth)

    pattern = excel_folder.format(year_pattern, month_pattern)
    return [f for f in glob.glob(pattern, recursive=False) if not os.path.basename(f).startswith('~')]

# Convert to Pandas DataFrame
def dataframe(header=None):
  def decorator(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        # Run the wrapped function, which should return (())/[[]]/generator
        data = f(*args, **kwargs)
        # Columns has to be passed to DataFrame Constructor since Columns can not be assigned to empty DataFrame Afterwards
        df = pd.DataFrame(data, columns=header if header else None)
        return df
    return decorated_function
  return decorator

# Reading A Range from Excel worksheet
def read_excel_data(file=None, sheet_name=0, range_address='A1:B2'):
  rows = ()
  if file:
    with open(file, 'rb') as f:
      in_mem_file = io.BytesIO(f.read())
      rng = openpyxl.load_workbook(
          in_mem_file, read_only=True, data_only=True)[sheet_name][range_address]
      # () is a generator comprehension (not tuple comprehension), include only the valid date(first value) row in excel
      rows = ((cell.value for cell in row)
              for row in rng if isinstance(row[0].value, datetime.datetime))
  return rows


daily_data_header = ('DateOpened', 'Category', 'PaymentType', 'NumberOfItems',
                     'TotalValueOfBatch', 'MailOpenedBy', 'CashAmount', 'CashTQBatch')

# Reading Daily Mail Data
@dataframe(daily_data_header)
def read_mail_daily_data(file=None):
  wbdata = []
  if file:
    with open(file, 'rb') as f:
      in_mem_file = io.BytesIO(f.read())
      workbook = openpyxl.load_workbook(
          in_mem_file, read_only=True, data_only=True)

      daily_data = {
          'pre201611': {
              'sheet_name_regex': '^[0-9][0-9]',
              'range_address_format': 'A{}:Q{}',
              'data_min_row': 8,
              'column_type': {
                  3: ('Donation', 'CashCheque'),
                  4: ('Donation', 'CashCheque'),
                  5: ('Donation', 'CashCheque'),
                  6: ('Donation', 'CreditCard'),
                  7: ('Donation', 'CreditCard'),
                  8: ('Donation', 'CreditCard'),
                  9: ('Merchandise', 'CashCheque'),
                  10: ('Merchandise', 'CashCheque'),
                  11: ('Merchandise', 'CreditCard'),
                  12: ('Merchandise', 'CreditCard'),
                  13: ('List', 'CashCheque'),
                  14: ('Other', 'CashCheque')
              }
          },
          'post201611': {
              'sheet_name_regex': '^[0-9][0-9]',
              'range_address_format': 'B{}:J{}',
              'data_min_row': 8,
              'column_type': {
                  1: ('Donation', 'CashCheque'),
                  2: ('Donation', 'CreditCard'),
                  3: ('Merchandise', 'CashCheque'),
                  4: ('Merchandise', 'CreditCard'),
                  5: ('List', 'CashCheque'),
                  6: ('Other', 'CashCheque'),
              }
          }
      }

      if str(os.path.splitext(file)[1]).endswith('sx'):
        # Pre 2016 11 Excels were marco enabled workbook
        sheet_info = daily_data['post201611']
        sorted_col_keys = sorted(list(sheet_info['column_type'].keys()))
        first_col, last_col = sorted_col_keys[0], sorted_col_keys[-1]

        vaild_worksheets = (wks for wks in workbook.worksheets if re.compile(
            sheet_info['sheet_name_regex']).match(wks.title))
        for wks in vaild_worksheets:
            # Read Workseet wild infomation: Date and Opener
            date_opened = wks.cell(row=1, column=3).value
            opener = wks.cell(row=1, column=5).value

            # Only the rows that has a total number of items are vaild rows
            range_address = sheet_info['range_address_format'].format(
                sheet_info['data_min_row'], wks.max_row)
            vaild_rows = (
                row for row in wks[range_address] if not row[0].value is None)

            for row in vaild_rows:
              for key, cell in enumerate(row):
                if not cell.value is None and first_col <= key <= last_col:
                  data = (date_opened, sheet_info['column_type'][key][0], sheet_info['column_type']
                          [key][1], row[0].value, cell.value, opener, row[-2].value, row[-1].value)
                  wbdata.append(data)
      else:
        # Post 2016 11 Excels are normal workbook
        sheet_info = daily_data['pre201611']
        sorted_col_keys = sorted(list(sheet_info['column_type'].keys()))
        first_col, last_col = sorted_col_keys[0], sorted_col_keys[-1]

        vaild_worksheets = (wks for wks in workbook.worksheets if re.compile(
            sheet_info['sheet_name_regex']).match(wks.title))
        for wks in vaild_worksheets:
            # Read Workseet wild infomation: Date and Opener
            date_opened = wks.cell(row=1, column=3).value
            opener = wks.cell(row=1, column=5).value

            # Only the rows that were not scraped are vaild rows
            range_address = sheet_info['range_address_format'].format(
                sheet_info['data_min_row'], wks.max_row)
            vaild_rows = (
                row for row in wks[range_address] if not row[0].value == True)

            for row in vaild_rows:
              for key, cell in enumerate(row):
                
                if not cell.value is None and first_col <= key <= last_col:
                  data = (date_opened, sheet_info['column_type'][key][0], sheet_info['column_type']
                          [key][1], row[2].value, cell.value, opener, None, None)
                  wbdata.append(data)
  return wbdata


all_files = find_mail_excels()
print('>>> {} Excel files found'.format(len(all_files)))

# Reading Individual Sheet for Daily Data then merge
mgdf = pd.concat([read_mail_daily_data(f)
                  for f in all_files], axis=0, ignore_index=True)

try:
    mgdf.to_csv(os.path.join(export_folder, 'MergedFromDetail.csv'),
                encoding='utf-8-sig')
    print('>>> Merged CSV File Saved [Using Daily Data]')
except PermissionError:
    print('Permission denied')

summary_sheet_name = 'SUMMARY'
summary_data_range = 'B3:S33'
summary_data_header = ('DateOpened', 'NumberOfItems', 'TotalValueOfBatch', 'MailOpenedBy',
                       'NoDCA', 'NoDCC', 'NoMCA', 'NoMCC', 'NoList', 'NoOther',
                       'ValDCA', 'ValDCC', 'ValMCA', 'ValMCC', 'ValList', 'ValOther',
                       'CashAmount', 'ValCashPending')

# Reading Summary Sheet for Monthly Data then merge
mgdf = pd.concat([pd.DataFrame(read_excel_data(f, summary_sheet_name, summary_data_range),
                               columns=summary_data_header) for f in all_files], axis=0, ignore_index=True)
try:
    mgdf.to_csv(os.path.join(export_folder,
                             'MergedFromSummary.csv'), encoding='utf-8-sig')
    print('>>> Merged CSV File Saved [Using Summary Data]')
except PermissionError:
    print('Permission denied')
